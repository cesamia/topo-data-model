"""Generate temp/10k_mapping_review.xlsx from changelog analysis."""
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path("temp/10k_mapping_review.xlsx")

# Status codes
CONFIRMED = "Confirmed"
FLAGGED   = "Flagged"
UNKNOWN   = "Needs Clarification"

# (10k_fc, geometry_10k, hdm_fc, geometry_hdm, status, notes)
MAPPINGS = [
    ("AdministrativeBoundary_A", "Polygon", "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("AdministrativeBoundary_L", "Line",    "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("AirportAirfield_A",        "Polygon", "RunwayA", "Polygon",CONFIRMED, ""),
    ("Alley_L",                  "Line",    "RoadL / TrailL", "Line", FLAGGED,
        "SPLIT: narrow passageways → RoadL (TUC=31); stairways → TrailL subtype 1 = StairsLine"),
    ("Bridge_L",                 "Line",    "BridgeL", "Line",   CONFIRMED, "Subtype 0 = BridgeOverpassViaductNodePoint"),
    ("Bridge_P",                 "Point",   "BridgeC", "Point",  CONFIRMED, "Subtype 0 = BridgeOverpassViaductNodePoint"),
    ("Building_A",               "Polygon", "BuildA",  "Polygon",CONFIRMED, "Subtype 0 = BuildingArea"),
    ("Building_P",               "Point",   "BuildP",  "Point",  CONFIRMED, "Subtype 0 = BuildingPoint"),
    ("Canal_L",                  "Line",    "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("Cemetery_A",               "Polygon", "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("Cemetery_P",               "Point",   "LandmrkP","Point",  CONFIRMED, "Subtype 2 = CemeteryPoint"),
    ("Cloud_A",                  "Polygon", "",        "",       UNKNOWN,   "No mapping in changelog (original alias was BndvoidA)"),
    ("Coast_A",                  "Polygon", "CoastA",  "Polygon",CONFIRMED, "Subtype 1 = WaterArea"),
    ("Coast_L",                  "Line",    "CoastL",  "Line",   CONFIRMED, "Subtype 0 = CoastlineShoreline"),
    ("CommunicationTower_P",     "Point",   "CommP",   "Point",  CONFIRMED, "Subtype 2 = CommunicationTowerPoint"),
    ("ComplexOutline_A",         "Polygon", "BndvoidA","Polygon",CONFIRMED, "Subtype 1 = ComplexOutlineArea"),
    ("Contour_L",                "Line",    "ContourL","Line",   CONFIRMED, ""),
    ("CoralReef_A",              "Polygon", "DangerA", "Polygon",CONFIRMED, "Subtype 1 = ReefArea"),
    ("CoralReef_L",              "Line",    "DangerL", "Line",   CONFIRMED, "Subtype 0 = ReefLine"),
    ("CoralReef_P",              "Point",   "",        "",       UNKNOWN,   "No mapping in changelog (possibly DangerP?)"),
    ("Crop_A",                   "Polygon", "CropA",   "Polygon",CONFIRMED, "Subtypes 0 = RiceFieldArea; 1 = CroplandArea"),
    ("Dam_L",                    "Line",    "DamL",    "Line",   CONFIRMED, "Subtype 0 = DamsWeirLine"),
    ("Ditch_A",                  "Polygon", "WatrcrsA","Polygon",CONFIRMED, "Subtype 1 = DitchArea"),
    ("Ditch_L",                  "Line",    "WatrcrsL","Line",   CONFIRMED, "Subtype 1 = DitchLine"),
    ("Fence_L",                  "Line",    "BarrierL","Line",   CONFIRMED, "Subtype 0 = FenceLine"),
    ("Ferry_L",                  "Line",    "FerryL",  "Line",   CONFIRMED, "Subtype 0 = FerryCrossingLine"),
    ("Fort_L",                   "Line",    "FortA",   "Polygon",FLAGGED,
        "GEOMETRY CHANGE: Line → Polygon. Subtype 0 = FortificationArea"),
    ("Grass_A",                  "Polygon", "GrassA",  "Polygon",CONFIRMED, "Subtypes 1 & 2 = GrasslandArea"),
    ("Grid",                     "Polygon", "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("Ground_A",                 "Polygon", "GroundA", "Polygon",CONFIRMED, "Subtype 0 = GroundSurfaceElementArea"),
    ("Heliport_P",               "Point",   "AerofacP","Point",  CONFIRMED, "Field apt = 9"),
    ("ICMHydro_P",               "Point",   "TlmhydroP","Point", CONFIRMED, "Subtype 2 = CurrentArrowPoint"),
    ("IndexMap",                 "Polygon", "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("LakePond_A",               "Polygon", "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("LakePond_L",               "Line",    "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("Lighthouse_P",             "Point",   "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("Markers_P",                "Point",   "MarkersP","Point",  CONFIRMED, "Subtype 2 = ControlPoint"),
    ("Military_P",               "Point",   "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("Mine_P",                   "Point",   "ExtractP","Point",  CONFIRMED, "Subtype 0 = MineQuarryPoint"),
    ("Misc_L",                   "Line",    "MiscL",   "Line",   CONFIRMED, "Subtype 0 = PenstockLine"),
    ("Monument_P",               "Point",   "LandmrkP","Point",  CONFIRMED, "Subtype 0 = MonumentPoint → remapped to subtype 3"),
    ("Mtn_P",                    "Point",   "MtnP",    "Point",  CONFIRMED, "Subtypes 0 = CavePoint; 1 = MountainPassPoint"),
    ("Orchard_A",                "Polygon", "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("Park_P",                   "Point",   "LandmrkP","Point",  CONFIRMED, "Subtype 3 = ParkPoint"),
    ("Pavement_A",               "Polygon", "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("Pier_A",                   "Polygon", "PierA",   "Polygon",CONFIRMED, ""),
    ("Pipe_L",                   "Line",    "PipeL",   "Line",   CONFIRMED, ""),
    ("PowerLine_L",              "Line",    "PowerL",  "Line",   CONFIRMED, "Subtype 0 = PowerTransmissionLine"),
    ("Quarry_A",                 "Polygon", "ExtractA","Polygon",CONFIRMED, "Subtype 0 = MineQuarryArea"),
    ("Quarry_L",                 "Line",    "",        "",       FLAGGED,
        "FLAG: 'Quarry_L' not found in changelog. Original 10k list had 'Quarry_P' — typo?"),
    ("Railroad_L",               "Line",    "RailrdL", "Line",   CONFIRMED, "Subtype 0 = RailroadTrackLine"),
    ("RiverStream_A",            "Polygon", "WatrcrsA","Polygon",CONFIRMED, "Subtype 2 = RiverStreamArea"),
    ("RiverStream_L",            "Line",    "WatrcrsL","Line",   CONFIRMED, "Subtype 2 = RiverStreamLine"),
    ("Road_A",                   "Polygon", "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("Road_L",                   "Line",    "RoadL",   "Line",   CONFIRMED, ""),
    ("RoadCasement_L",           "Line",    "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("RoadMedian_A",             "Polygon", "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("Ruins_A",                  "Polygon", "RuinsA",  "Polygon",CONFIRMED, ""),
    ("SideWalk_L",               "Line",    "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("Slope_A",                  "Polygon", "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("Sport_A",                  "Polygon", "SportA",  "Polygon",CONFIRMED, ""),
    ("SpotElevation_P",          "Point",   "ElevP",   "Point",  CONFIRMED, "Subtype 0 = SpotElevationPoint"),
    ("Swamp_A",                  "Polygon", "SwampA",  "Polygon",CONFIRMED,
        "Subtype 1 = MarshSwampArea carried over; subtype 0 unmapped (removed)"),
    ("Tank_A",                   "Polygon", "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("Tank_P",                   "Point",   "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("Tower_P",                  "Point",   "TlmutilP","Point",  CONFIRMED,
        "Subtype 0 = TowerNonCommunicationPoint → PowerTransmissionPylonPoint"),
    ("Trail_L",                  "Line",    "",        "",       UNKNOWN,   "No mapping in changelog (possibly TrailL?)"),
    ("Trans_L",                  "Line",    "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("Treat_A",                  "Polygon", "TreatA",  "Polygon",CONFIRMED, "Subtypes 0, 1, 2, 3 mapped"),
    ("Treat_L",                  "Line",    "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("Trees_A",                  "Polygon", "TreesA",  "Polygon",CONFIRMED, ""),
    ("Trees_P",                  "Point",   "TreesP",  "Point",  CONFIRMED, ""),
    ("Tunnel_L",                 "Line",    "TunnelL", "Line",   CONFIRMED, ""),
    ("Veg_A",                    "Polygon", "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("Veg_L",                    "Line",    "",        "",       UNKNOWN,   "No mapping in changelog"),
    ("Wellspr_P",                "Point",   "WellsprP","Point",  CONFIRMED, "Subtypes 0 & 1 mapped"),
]

assert len(MAPPINGS) == 76, f"Expected 76, got {len(MAPPINGS)}"

# ── Styles ────────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1A3A5C")
HDR_FONT   = Font(color="FFFFFF", bold=True, size=10)
CONF_FILL  = PatternFill("solid", fgColor="E8F5EE")
FLAG_FILL  = PatternFill("solid", fgColor="FEF9E8")
UNK_FILL   = PatternFill("solid", fgColor="FDE8E8")
WRAP       = Alignment(wrap_text=True, vertical="top")
THIN       = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

STATUS_FILL = {CONFIRMED: CONF_FILL, FLAGGED: FLAG_FILL, UNKNOWN: UNK_FILL}
STATUS_COLOR = {CONFIRMED: "1E7E48", FLAGGED: "B7860D", UNKNOWN: "C0392B"}


def style_cell(cell, fill=None, font_color=None, bold=False, size=10):
    if fill:
        cell.fill = fill
    cell.font = Font(color=font_color or "2C3E50", bold=bold, size=size)
    cell.alignment = WRAP
    cell.border = THIN


def build_mapping_sheet(wb):
    ws = wb.active
    ws.title = "10k FC Mapping"

    headers = [
        "10k FC Name", "Geometry (10k)",
        "HDM FC (Mapped To)", "Geometry (HDM)",
        "Status", "Notes / Subtype Detail",
        "Confirmed By",   # blank — for user to fill
        "Correction / Override",  # blank — for user to fill
    ]
    col_widths = [28, 14, 22, 14, 20, 52, 18, 30]

    # Header row
    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 30

    # Data rows
    for r, (fc, geo10, hdm, geohdm, status, notes) in enumerate(MAPPINGS, 2):
        row_fill = STATUS_FILL[status]
        row_data = [fc, geo10, hdm, geohdm, status, notes, "", ""]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = row_fill
            cell.alignment = WRAP
            cell.border = THIN
            if c == 1:
                cell.font = Font(bold=True, size=10, color="2C3E50")
            elif c == 5:
                cell.font = Font(bold=True, size=10, color=STATUS_COLOR[status])
            else:
                cell.font = Font(size=10, color="2C3E50")
        ws.row_dimensions[r].height = 30 if not notes else max(30, min(60, len(notes) // 2))

    # Freeze header
    ws.freeze_panes = "A2"

    # Summary below
    last = len(MAPPINGS) + 3
    counts = {s: sum(1 for m in MAPPINGS if m[4] == s) for s in (CONFIRMED, FLAGGED, UNKNOWN)}
    ws.cell(row=last, column=1, value="Summary").font = Font(bold=True, size=10)
    for i, (s, n) in enumerate(counts.items(), 1):
        ws.cell(row=last + i, column=1, value=s).font = Font(
            bold=True, size=10, color=STATUS_COLOR[s]
        )
        ws.cell(row=last + i, column=2, value=n).font = Font(size=10)


def build_flags_sheet(wb):
    ws = wb.create_sheet("Flags & Inconsistencies")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70

    flags = [
        ("SPLIT mapping",
         "Alley_L (10k) maps to TWO HDM FCs: RoadL (narrow passageways, TUC=31) "
         "and TrailL subtype 1 = StairsLine. Sidebar will link to RoadL by default — confirm preference."),
        ("Geometry change",
         "Fort_L (10k, Line) maps to FortA (HDM, Polygon). "
         "This is a geometry type change documented in the changelog: 'Note change in geometry type.'"),
        ("Quarry_L not in changelog",
         "Your 10k list contains 'Quarry_L'. The changelog only documents 'Quarry_A → ExtractA'. "
         "The original 10k HTML list had 'Quarry_P', not 'Quarry_L'. Possible typo — please confirm."),
        ("GeoNames_P absent from new 10k list",
         "GeoNames_P was in the original 10k HTML list. The changelog maps it to: NamedLocP "
         "(primary), plus partial mappings to RapidsC, ThermalP, HarborP, and PowerP. "
         "It is NOT in your new 76-FC list. Was this intentional?"),
        ("PumpingA / PumpingP (50k)",
         "Changelog: 'Deleted from 50k with 10k equivalent, but different container.' "
         "Building_A (10k) → BuildA (HDM), NOT PumpingA. "
         "Both PumpingA and PumpingP show correctly as DEL in the 50k sidebar."),
        ("Multiple 10k FCs → single HDM FC (merges)",
         "Several merges occurred:\n"
         "  • Ditch_A + RiverStream_A → WatrcrsA\n"
         "  • Ditch_L + RiverStream_L → WatrcrsL\n"
         "  • Cemetery_P + Monument_P + Park_P → LandmrkP\n"
         "All three 10k sidebar items will link to the same HDM page — verify this is the intended behaviour."),
        ("Swamp_A partial mapping",
         "Swamp_A subtype 1 = MarshSwampArea is mapped to SwampA. "
         "Subtype 0 has no HDM equivalent (removed). The changelog does not clarify what happens to subtype 0 data."),
    ]

    hdr_row = ["Flag / Issue", "Detail"]
    for c, h in enumerate(hdr_row, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN

    for r, (flag, detail) in enumerate(flags, 2):
        c1 = ws.cell(row=r, column=1, value=flag)
        c1.fill = FLAG_FILL
        c1.font = Font(bold=True, size=10, color=STATUS_COLOR[FLAGGED])
        c1.alignment = WRAP
        c1.border = THIN

        c2 = ws.cell(row=r, column=2, value=detail)
        c2.fill = FLAG_FILL
        c2.font = Font(size=10, color="2C3E50")
        c2.alignment = WRAP
        c2.border = THIN
        ws.row_dimensions[r].height = max(40, len(detail) // 1)

    ws.freeze_panes = "A2"


# ── Main ──────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
build_mapping_sheet(wb)
build_flags_sheet(wb)
wb.save(OUT)
print(f"Written -> {OUT}  ({len(MAPPINGS)} rows)")
print(f"  Confirmed : {sum(1 for m in MAPPINGS if m[4]==CONFIRMED)}")
print(f"  Flagged   : {sum(1 for m in MAPPINGS if m[4]==FLAGGED)}")
print(f"  Needs input: {sum(1 for m in MAPPINGS if m[4]==UNKNOWN)}")
